import openpyxl
import xlrd


def account():
    print("###### WELCOME TO BAJRANGDAS CENTRAL LIBRARY ######")
    r = input("\nDo you have an existing account?(y/n)\n")
    loc = "C:\\SDP Project\\Details.xlsx"
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)
    loc2 = "C:\\SDP Project\\Books.xlsx"
    wb2 = xlrd.open_workbook(loc2)
    sheet2 = wb2.sheet_by_index(0)
    sheet2.cell_value(0, 0)
    xfile = openpyxl.load_workbook("Details.xlsx")
    sheetd = xfile.active
    n = sheetd.max_row
    flag = 0
    if r == "y" or r == "Y":
        user = input("Enter Your Username")
        passw = input("Enter Your Password")
        for i in range(sheet.nrows):
            if sheet.cell_value(i, 2) == user and sheet.cell_value(i, 3) == passw:
                """Comment the next 2 lines and add function call when the function issue() is created"""
                # for k in range(sheet2.nrows):
                #         print (sheet2.cell_value(k, 0),"  ",sheet2.cell_value(k, 1))
                issue(user)
                flag = 1
            if flag == 0:
                print("INCORRECT USERNAME OR PASSWORD")
            elif r == "n" or r == "N":
                f = 0
            while f != 1:
                name = input("Enter Your Name\n")
                gr = int(input("Enter Your GR Number\n"))
                user = input("Enter Your Username")
                passw = input("Enter Your Password")
                for i in range(sheet.nrows):
                    if sheet.cell_value(i, 2) == user or sheet.cell_value(i, 1) == gr:
                        print(
                            "USERNAME ALREADY TAKEN  or Duplicate GR Number\nTRY AGAIN"
                        )
                        f = 0
                        break
                    else:
                        f = 1
                        c1 = sheetd.cell(n + 1, 1)
                        c1.value = name
                        c2 = sheetd.cell(n + 1, 2)
                        c2.value = gr
                        c3 = sheetd.cell(n + 1, 3)
                        c3.value = user
                        c4 = sheetd.cell(n + 1, 4)
                        c4.value = passw
                        xfile.save(loc)
                        """Comment the next 2 lines and add function call when the function issue() is created """
                # for k in range(sheet2.nrows):
                #          print (sheet2.cell_value(k, 0), "  ", sheet2.cell_value(k, 1))
        issue(user)


def issue(usern):
    loc2 = "C:\\SDP Project\\Books.xlsx"
    wb2 = xlrd.open_workbook(loc2)
    sheet2 = wb2.sheet_by_index(0)
    sheet2.cell_value(0, 0)
    loc = "C:\\SDP Project\\Details.xlsx"
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)
    xfile = openpyxl.load_workbook("Details.xlsx")
    sheetd = xfile.active
    xfile2 = openpyxl.load_workbook("Books.xlsx")
    sheetb = xfile2.active
    ch = "y"
    # print (n)
    for m in range(100):
        option = int(
            input(
                "Enter\n1.To View Available Books\n2.To View the books issued by you\n3.Return Books \n4.Login Again\n0.To Exit"
            )
        )
        if option == 1:
            for k in range(sheet2.nrows):
                print(sheet2.cell_value(k, 0), "  ", sheet2.cell_value(k, 1))
                while ch == "y":
                    serial = int(
                        input("Enter the Serial Number of the book you want to issue:")
                    )
                    # print (ch)
                    if sheet2.cell_value(serial + 1, 5) == 0:
                        print("No Copies Left\n")
                        continue
                    else:
                        for k in range(sheet2.nrows):
                            if sheet2.cell_value(k, 0) == serial:
                                for i in range(sheet.nrows):
                                    if sheet.cell_value(i, 2) == usern:
                                        # n = sheetd.max_column
                                        c1 = sheetd.cell(i + 1, len(sheetd[i + 1]) + 1)
                                        c1.value = sheet2.cell_value(k, 1)
                                        c2 = sheetb.cell(serial + 2, 6)
                                        c2.value = sheet2.cell_value(serial + 1, 5) - 1
                                        n = sheetd.max_column
                                        xfile.save(loc)
                                        xfile2.save(loc2)
                                        print(
                                            "\n\n\nSUCCESSFULL !!!!!\n The changes will be reflected in your account \n When you Login Again"
                                        )
                                        ch = input("Do You want to continue?(y/n)\n")
                                        if ch == "n" or ch == "N":
                                            break
                                        elif option == 3:
                                            cnt = 1
                                        print("The books issued by you are:\n")
                                        for i in range(sheet.nrows):
                                            if sheet.cell_value(i, 2) == usern:
                                                num = i
                                            for j in range(4, sheet.ncols):
                                                if (
                                                    sheetd.cell(i + 1, j + 1).value
                                                    is not None
                                                ):
                                                    print(
                                                        cnt,
                                                        ".",
                                                        sheet.cell_value(i, j),
                                                        end="\n",
                                                    )
                                                    cnt = cnt + 1
                                                    retch = "y"
                                                    for i in range(sheet.ncols - 4):
                                                        ret = int(
                                                            input(
                                                                "Enter the number You want to return:"
                                                            )
                                                        )
                                                    for k in range(sheet2.nrows):
                                                        for j in range(
                                                            4, len(sheetd[num + 1])
                                                        ):
                                                            if (
                                                                sheet.cell_value(
                                                                    num, 4 + ret - 1
                                                                )
                                                                != ""
                                                            ):
                                                                break
                                                            else:
                                                                ret = ret + 1
                                                            if sheet2.cell_value(
                                                                k, 1
                                                            ) == sheet.cell_value(
                                                                num, 4 + ret - 1
                                                            ):
                                                                c = sheetb.cell(
                                                                    k + 1, 6
                                                                )
                                                                c.value = (
                                                                    sheet2.cell_value(
                                                                        k, 5
                                                                    )
                                                                    + 1
                                                                )
                                                                # sheetd .delete_cells(num+1,ret+4)
                                                                sheetd.cell(
                                                                    num + 1, ret + 4
                                                                ).value = None
                                                                print(
                                                                    "\nReturn successful"
                                                                )
                                                                xfile.save(loc)
                                                                xfile2.save(loc2)
                                                                retch = input(
                                                                    "Do you have any books to return ?(y/n)"
                                                                )
                                                            if (
                                                                retch == "n"
                                                                or retch == "N"
                                                            ):
                                                                break
                                                            elif (
                                                                retch == "y"
                                                                or retch == "Y"
                                                            ):
                                                                print(
                                                                    "\n\n\nCONFIRM YOUR LOGIN DETAILS\n"
                                                                )
                                                                account()
                                                            elif option == 0:
                                                                print(
                                                                    "###### Thank You ######\n\n###### Visit Again #####"
                                                                )
                                                                break
                                                            elif option == 4:
                                                                account()
                                                            elif option == 2:
                                                                cnt = 1
                                                                print(
                                                                    "The books issued by you are:\n"
                                                                )
                                                            for i in range(sheet.nrows):
                                                                if (
                                                                    sheet.cell_value(
                                                                        i, 2
                                                                    )
                                                                    == usern
                                                                ):
                                                                    num = i
                                                                for j in range(
                                                                    4, sheet.ncols
                                                                ):
                                                                    if (
                                                                        sheetd.cell(
                                                                            i + 1, j + 1
                                                                        ).value
                                                                        is not None
                                                                    ):
                                                                        print(
                                                                            cnt,
                                                                            ".",
                                                                            sheet.cell_value(
                                                                                i, j
                                                                            ),
                                                                            end="\n",
                                                                        )
                                                                        cnt = cnt + 1


account()
