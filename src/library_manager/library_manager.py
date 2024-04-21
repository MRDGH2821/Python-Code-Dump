"""Library Manager.

This program is a library manager that allows users to issue and return books.
"""

from __future__ import annotations

import os
import sys

from openpyxl import load_workbook

current_directory = os.path.dirname(os.path.abspath(__file__))


details_file_path = f"{current_directory}/Details.xlsx"
books_file_path = f"{current_directory}/Books.xlsx"


def find_user(username: str):
    """Find a user by username."""
    details_workbook = load_workbook(details_file_path)
    details_sheet = details_workbook[details_workbook.sheetnames[0]]

    for i in details_sheet.iter_rows(min_row=1):
        username_value = i[2].value
        if username_value == username:
            return i
    return None


def account() -> None:
    """Menu of program."""
    print("###### WELCOME TO BAJRANGDAS CENTRAL LIBRARY ######")
    choice = input("\nDo you have an existing account?(y/n): ")
    details_workbook = load_workbook(details_file_path)
    details_sheet = details_workbook[details_workbook.sheetnames[0]]

    if choice in ("y", "Y"):
        print("\nUser login")
        username = input("Enter Your Username: ")
        password = input("Enter Your Password: ")
        i = find_user(username)
        if i:
            username_value = i[2].value
            password_value = i[3].value
            # print(username_value, password_value)
            if username_value == username and password_value == password:
                issue(username)
            else:
                msg = "Invalid Username or Password."
                raise ValueError(msg)
        else:
            msg = "Unknown Username or Password. Please register."
            raise ValueError(msg)

    elif choice in ("n", "N"):
        print("\nUser Registration")
        name = input("Enter Your Name: ")
        gr_number = int(input("Enter Your GR Number: "))
        username = input("Enter Your Username: ")
        password = input("Enter Your Password: ")

        i = find_user(username)
        if i:
            gr_number_value = i[1].value
            username_value = i[2].value
            password_value = i[3].value

            if username_value == username or gr_number_value == gr_number:
                msg = "USERNAME ALREADY TAKEN or Duplicate GR Number\nTRY AGAIN"
                raise ValueError(msg)

        else:
            details_sheet.append([name, gr_number, username, password])
            details_workbook.save(details_file_path)

        issue(username)


def issue(username: str):
    """Issue a book."""
    books_workbook = load_workbook(books_file_path)
    books_sheet = books_workbook[books_workbook.sheetnames[0]]

    details_workbook = load_workbook(details_file_path)
    details_sheet = details_workbook[details_workbook.sheetnames[0]]

    choice = "y"

    while True:
        print("\n\nMain Menu:")
        print("1. View Available Books")
        print("2. View the books issued by you")
        print("3. Return Books")
        print("0. Exit (default)")
        option = int(input("Enter choice: "))
        if option == 1:
            print("Available books:\n")
            for row in books_sheet.iter_rows():
                print(row[0].value, "  ", row[1].value)
            while choice == "y":
                serial = int(
                    input("Enter the Serial Number of the book you want to issue:"),
                )
                # print (ch)
                if books_sheet.cell(serial + 1, 5).value == 0:
                    print("No Copies Left\n")
                    continue
                else:
                    for k in range(1, books_sheet.max_row):
                        if books_sheet.cell(k, 0).value == serial:
                            for i in range(1, details_sheet.max_row):
                                if details_sheet.cell(i, 2).value == username:
                                    # n = sheetd.max_column
                                    c1 = details_sheet.cell(
                                        i + 1,
                                        len(details_sheet[i + 1]) + 1,
                                    )
                                    c1.value = books_sheet.cell(k, 1).value
                                    c2 = books_sheet.cell(serial + 2, 6)
                                    c2.value = books_sheet.cell(serial + 1, 5).value - 1
                                    details_workbook.save(details_file_path)
                                    books_workbook.save(books_file_path)
                                    print(
                                        "\n\n\nSUCCESSFULL !!!!!\n The changes will be reflected in your account \n When you Login Again",
                                    )
                                    choice = input("Do You want to continue?(y/n)\n")
                                    if choice in ("n", "N"):
                                        break
        elif option == 2:
            count = 0
            print("The books issued by you are:\n")
            for i in range(1, details_sheet.max_row):
                if details_sheet.cell(i, 2).value == username:
                    num = i
                for j in range(4, details_sheet.max_column):
                    if details_sheet.cell(i + 1, j + 1).value is not None:
                        print(
                            count,
                            ".",
                            details_sheet.cell(i, j).value,
                            end="\n",
                        )
                        count = count + 1
        elif option == 3:
            count = 0
            print("The books issued by you are:\n")
            num = 0
            for i in range(1, details_sheet.max_row):
                if details_sheet.cell(i, 2).value == username:
                    num = i
                for j in range(4, details_sheet.max_column):
                    if details_sheet.cell(i + 1, j + 1).value is not None:
                        print(
                            count,
                            ".",
                            details_sheet.cell(i, j).value,
                            end="\n",
                        )
                        count = count + 1
                        return_choice = "y"
                        for _ in range(1, details_sheet.max_column - 4):
                            return_qty = int(
                                input("Enter the number You want to return:"),
                            )
                            for k in range(1, books_sheet.max_row):
                                for _j in range(4, len(details_sheet[num + 1])):
                                    if (
                                        details_sheet.cell(
                                            num,
                                            4 + return_qty - 1,
                                        ).value
                                        != ""
                                    ):
                                        break
                                    else:
                                        return_qty = return_qty + 1
                                    if (
                                        books_sheet.cell(k, 1).value
                                        == details_sheet.cell(
                                            num,
                                            4 + return_qty - 1,
                                        ).value
                                    ):
                                        c = books_sheet.cell(k + 1, 6)
                                        c.value = int(books_sheet.cell(k, 5).value) + 1
                                        # sheetd .delete_cells(num+1,ret+4)
                                        details_sheet.cell(
                                            num + 1,
                                            return_qty + 4,
                                        ).value = None
                                        print("\nReturn successful")
                                        details_workbook.save(details_file_path)
                                        books_workbook.save(books_file_path)
                                        return_choice = input(
                                            "Do you have any books to return ?(y/n)",
                                        )
                                    if return_choice in ("n", "N"):
                                        break
                                    elif return_choice in ("y", "Y"):
                                        print("\n\n\nCONFIRM YOUR LOGIN DETAILS\n")
                                        account()
        else:
            print("###### Thank You ######\n\n###### Visit Again #####")
            sys.exit(0)


account()
